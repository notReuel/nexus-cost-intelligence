"""
BOQ File Parser — Priority 3 (file handling layer)

Reads an XLSX or CSV BOQ file. Auto-detects the header row and extracts
line items into a normalized structure.

Returns: { 'sheet': str, 'header_row': int, 'columns': {...}, 'rows': [...] }

The output 'rows' is a list of dicts with at minimum:
  { 'description': str, 'qty': float?, 'unit': str?, 'rate': float?, 'amount': float?, 'raw_row': [...] }
"""
import csv
import io
import re
from openpyxl import load_workbook
from typing import Optional


# Column header dictionaries — match flexibly
HEADER_PATTERNS = {
    'description': re.compile(r'\b(description|item|particulars?|service|scope|details?|specification)\b', re.IGNORECASE),
    'qty':         re.compile(r'\b(qty|quantity|qnty|nos?|no\.)\b', re.IGNORECASE),
    'unit':        re.compile(r'\b(unit|uom|u\.o\.m|measure)\b', re.IGNORECASE),
    'rate':        re.compile(r'\b(rate|unit\s*price|unit\s*cost|price\s*per|usd\s*\/|\$\s*\/)\b', re.IGNORECASE),
    'amount':      re.compile(r'\b(amount|total|sub\s*total|extended|line\s*total)\b', re.IGNORECASE),
    's_n':         re.compile(r'\b(s\.?\s*n\.?|sl\.?\s*no\.?|sr\.?\s*no\.?|serial|item\s*no\.?|#)\b', re.IGNORECASE),
}


def _find_header_row(rows, max_scan=30):
    """Return (header_row_index, column_map) by scanning the first N rows for
    the row that contains the most matching header tokens.
    """
    best_row = -1
    best_map = {}
    best_score = 0
    for ri, row in enumerate(rows[:max_scan]):
        col_map = {}
        for ci, cell in enumerate(row):
            if cell is None: continue
            cell_str = str(cell).strip()
            if not cell_str or len(cell_str) > 60: continue
            for field, pat in HEADER_PATTERNS.items():
                if pat.search(cell_str) and field not in col_map:
                    col_map[field] = ci
                    break
        if 'description' in col_map and len(col_map) > best_score:
            best_row = ri
            best_map = col_map
            best_score = len(col_map)
    return best_row, best_map


def _to_float(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)): return float(v)
    s = str(v).replace(',', '').replace('₦', '').replace('$', '').strip()
    s = re.sub(r'[^\d.\-]', '', s)
    try: return float(s)
    except ValueError: return None


def _to_str(v):
    return str(v).strip() if v is not None else ''


def parse_xlsx(path: str, sheet_name: Optional[str] = None):
    """Parse an XLSX BOQ. Returns the parse result, including auto-detected
    headers and extracted line items. If sheet_name is None, picks the first
    sheet that has a recognisable BOQ header."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets_to_try = [sheet_name] if sheet_name else wb.sheetnames

    for sn in sheets_to_try:
        if sn not in wb.sheetnames: continue
        ws = wb[sn]
        # Read first 60 rows into memory for header detection
        rows = []
        for row_cells in ws.iter_rows(min_row=1, max_row=60, values_only=True):
            rows.append(list(row_cells))
        # If sheet has more rows, also scan beyond row 60 for actual content
        header_row, col_map = _find_header_row(rows)
        if header_row < 0 or 'description' not in col_map:
            continue  # try next sheet

        # Now extract all data rows after the header
        line_items = []
        for row_cells in ws.iter_rows(min_row=header_row + 2, values_only=True):
            row = list(row_cells)
            if not row or all(c is None or c == '' for c in row):
                continue
            desc_idx = col_map.get('description')
            if desc_idx is None or desc_idx >= len(row): continue
            desc = _to_str(row[desc_idx])
            if not desc or len(desc) < 3: continue
            # Skip subtotal / total rows
            if re.search(r'\b(sub\s*total|grand\s*total|total)\b', desc, re.IGNORECASE):
                continue
            li = {
                'description': desc,
                'qty':    _to_float(row[col_map['qty']])    if col_map.get('qty')    is not None and col_map['qty']    < len(row) else None,
                'unit':   _to_str  (row[col_map['unit']])   if col_map.get('unit')   is not None and col_map['unit']   < len(row) else None,
                'rate':   _to_float(row[col_map['rate']])   if col_map.get('rate')   is not None and col_map['rate']   < len(row) else None,
                'amount': _to_float(row[col_map['amount']]) if col_map.get('amount') is not None and col_map['amount'] < len(row) else None,
                'raw_row': [_to_str(c) for c in row],
            }
            # Derive rate from amount/qty if missing
            if li['rate'] is None and li['amount'] and li['qty'] and li['qty'] > 0:
                li['rate'] = li['amount'] / li['qty']
            # Derive amount if missing
            if li['amount'] is None and li['rate'] and li['qty']:
                li['amount'] = li['rate'] * li['qty']
            line_items.append(li)

        return {
            'sheet': sn,
            'header_row': header_row + 1,  # 1-indexed for display
            'columns': col_map,
            'rows': line_items,
            'total_rows': len(line_items),
        }
    return None  # no sheet matched


def parse_csv(path_or_text: str, is_text: bool = False):
    """Parse a CSV BOQ file or string."""
    if is_text:
        text = path_or_text
    else:
        with open(path_or_text, 'r', encoding='utf-8-sig', errors='replace') as f:
            text = f.read()

    reader = csv.reader(io.StringIO(text))
    rows = [r for r in reader]
    header_row, col_map = _find_header_row(rows)
    if header_row < 0 or 'description' not in col_map:
        return None

    line_items = []
    for row in rows[header_row + 1:]:
        if not row or all(c == '' for c in row): continue
        desc_idx = col_map.get('description')
        if desc_idx is None or desc_idx >= len(row): continue
        desc = _to_str(row[desc_idx])
        if not desc or len(desc) < 3: continue
        if re.search(r'\b(sub\s*total|grand\s*total|total)\b', desc, re.IGNORECASE):
            continue
        li = {
            'description': desc,
            'qty':    _to_float(row[col_map['qty']])    if col_map.get('qty')    is not None and col_map['qty']    < len(row) else None,
            'unit':   _to_str  (row[col_map['unit']])   if col_map.get('unit')   is not None and col_map['unit']   < len(row) else None,
            'rate':   _to_float(row[col_map['rate']])   if col_map.get('rate')   is not None and col_map['rate']   < len(row) else None,
            'amount': _to_float(row[col_map['amount']]) if col_map.get('amount') is not None and col_map['amount'] < len(row) else None,
            'raw_row': [_to_str(c) for c in row],
        }
        if li['rate'] is None and li['amount'] and li['qty'] and li['qty'] > 0:
            li['rate'] = li['amount'] / li['qty']
        if li['amount'] is None and li['rate'] and li['qty']:
            li['amount'] = li['rate'] * li['qty']
        line_items.append(li)

    return {
        'sheet': 'CSV',
        'header_row': header_row + 1,
        'columns': col_map,
        'rows': line_items,
        'total_rows': len(line_items),
    }
